# Copyright (c) 2025, Finbyz Tech Pvt Ltd and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe import whitelist
import os
from typing import List, Dict, Any

from langchain.text_splitter import RecursiveCharacterTextSplitter
from finbyzai.ai.vectorstores.registry import create_vector_store
from finbyzai.ai.embeddings.registry import create_embedding


class KnowledgeBase(Document):
	"""
	KnowledgeBase DocType that integrates with different vector stores
	via the registry + factory pattern with intelligent text chunking.
	"""
	def autoname(self):
		self.name = frappe.scrub(self.title)
	
	def get_text_splitter(self):
		"""
		Get configured text splitter with optimal chunking parameters.
		
		Recommended chunk sizes:
		- OpenAI embeddings: 512-1000 tokens (typically 2000-4000 chars)
		- Cohere embeddings: 512 tokens (typically 2000 chars)
		- For general use: 1000 chars with 200 overlap works well
		"""
		chunk_size = self.chunk_size or 1000
		chunk_overlap = self.chunk_overlap or 200
		
		# Use RecursiveCharacterTextSplitter as it's the recommended default
		return RecursiveCharacterTextSplitter(
			chunk_size=chunk_size,
			chunk_overlap=chunk_overlap,
			length_function=len,
			separators=["\n\n", "\n", ". ", " ", ""],
			is_separator_regex=False,
		)
		
	def get_vector_store(self):
		"""Return a vector store instance based on this KB's config."""
		store_name = (self.vector_store or "").strip().lower()
		if not store_name:
			return None

		emb = _get_embeddings(self)
		if not emb:
			raise frappe.ValidationError("Embeddings not configured")

		api_key = _get_provider_api_key(self)

		return create_vector_store(
			store_name=store_name,
			kb_name=self.name,
			description=self.description or "",
			embeddings=emb,
			api_key=api_key,
		)

	@whitelist()
	def upsert(self) -> int:
		"""
		Upsert all documents in this KB into the configured vector store.
		Now with intelligent text chunking for better retrieval.
		"""
		store = self.get_vector_store()
		if not store:
			raise frappe.ValidationError("No vector_store selected on Knowledge Base")

		text_splitter = self.get_text_splitter()
		items = []
		
		for row in self.documents or []:
			if row.is_process:
				continue
			if not row.text_content:
				continue
				
			# Split the text into chunks
			chunks = text_splitter.split_text(row.text_content)
			
			# Create separate items for each chunk with enhanced metadata
			for chunk_idx, chunk_text in enumerate(chunks):
				chunk_id = f"{self.name}-{row.name}-chunk-{chunk_idx}"
				
				# Enhanced metadata for better filtering and retrieval
				metadata = {
					"kb": self.name,
					"row_name": row.name,
					"file": row.file or "",
					"chunk_index": chunk_idx,
					"total_chunks": len(chunks),
					"source_type": self._get_source_type(row.file),
					# Add any custom metadata from the row
					"title": getattr(row, 'title', '') or '',
					"created_on": str(row.creation) if hasattr(row, 'creation') else '',
				}
				
				items.append({
					"id": chunk_id,
					"text": chunk_text,
					"metadata": metadata,
				})
			
			# Mark as processed after chunking
			row.is_process = 1
		
		if not items:
			frappe.msgprint("No new documents to process.", alert=True, indicator="blue")
			return 0
		
		self.save()
		
		# Upsert chunks to vector store
		num_upserted = store.upsert(
			texts=[it["text"] for it in items],
			metadatas=[it["metadata"] for it in items],
			ids=[it["id"] for it in items],
		)
		
		frappe.msgprint(
			f"Successfully processed {len(self.documents)} documents into {len(items)} chunks.",
			alert=True,
			indicator="green"
		)
		return num_upserted

	def _get_source_type(self, file_path: str) -> str:
		"""Determine the source type from file extension."""
		if not file_path:
			return "text"
		ext = file_path.lower().split('.')[-1]
		return {
			'pdf': 'pdf',
			'xlsx': 'excel',
			'xls': 'excel',
			'csv': 'csv',
			'txt': 'text',
			'md': 'markdown',
		}.get(ext, 'unknown')

	@whitelist()
	def search(self, query: str, limit: int = 5, filter_dict: dict = None) -> List[Dict[str, Any]]:
		"""
		Search this KB via vector store with optional metadata filtering.
		
		Args:
			query: Search query string
			limit: Number of results to return
			filter_dict: Optional metadata filters (e.g., {"source_type": "pdf"})
		"""
		store = self.get_vector_store()
		if not store:
			return self.simple_search(query, limit)

		# If your vector store supports metadata filtering, pass it through
		if filter_dict and hasattr(store, 'similarity_search'):
			return store.similarity_search(query, k=limit, filter=filter_dict)
		
		return store.search(query, k=limit)

	def simple_search(self, query: str, limit: int = 5) -> List[Dict[str, Any]]:
		"""Fallback keyword search if no vector store configured."""
		results: List[Dict[str, Any]] = []
		if not query or not self.documents:
			return results

		needle = query.strip().lower()
		for row in self.documents:
			text = (row.text_content or "").lower()
			if needle in text:
				idx = text.find(needle)
				snippet = row.text_content[max(0, idx - 60): idx + len(needle) + 60]
				results.append({
					"id": row.name,
					"score": text.count(needle),
					"snippet": snippet,
					"file": row.file or "",
				})

		return sorted(results, key=lambda r: r["score"], reverse=True)[:limit]

	@whitelist()
	def search_with_context(self, query: str, limit: int = 5, context_chars: int = 200) -> List[Dict[str, Any]]:
		"""
		Enhanced search that returns results with surrounding context.
		
		Args:
			query: Search query string
			limit: Number of results to return
			context_chars: Number of characters to show before/after match
		
		Returns:
			List of search results with context snippets
		"""
		store = self.get_vector_store()
		if not store:
			return self.simple_search_with_context(query, limit, context_chars)

		# Get vector search results
		raw_results = store.search(query, k=limit)
		
		# Enhance with context
		enhanced = []
		for result in raw_results:
			metadata = result.get("metadata", {})
			row_name = metadata.get("row_name")
			chunk_idx = metadata.get("chunk_index", 0)
			
			# Find the document
			doc_text = ""
			source_file = ""
			for doc in self.documents:
				if doc.name == row_name:
					doc_text = doc.text_content or ""
					source_file = doc.file.split('/')[-1] if doc.file else "Text Document"
					break
			
			# Get the specific chunk
			if doc_text:
				text_splitter = self.get_text_splitter()
				chunks = text_splitter.split_text(doc_text)
				
				snippet = chunks[chunk_idx] if chunk_idx < len(chunks) else doc_text[:500]
				
				enhanced.append({
					"id": result.get("id"),
					"score": result.get("score"),
					"snippet": snippet,
					"source": source_file,
					"chunk_index": chunk_idx,
					"total_chunks": len(chunks),
					"metadata": metadata
				})
		
		return enhanced

	def simple_search_with_context(self, query: str, limit: int = 5, context_chars: int = 200) -> List[Dict[str, Any]]:
		"""Fallback keyword search with context if no vector store configured."""
		results = []
		if not query or not self.documents:
			return results

		needle = query.strip().lower()
		for row in self.documents:
			text = (row.text_content or "").lower()
			original_text = row.text_content or ""
			
			if needle in text:
				idx = text.find(needle)
				start = max(0, idx - context_chars)
				end = min(len(text), idx + len(needle) + context_chars)
				
				snippet = original_text[start:end]
				if start > 0:
					snippet = "..." + snippet
				if end < len(text):
					snippet = snippet + "..."
				
				results.append({
					"id": row.name,
					"score": text.count(needle),
					"snippet": snippet,
					"source": row.file.split('/')[-1] if row.file else "Text Document",
					"metadata": {
						"row_name": row.name,
						"kb": self.name,
						"file": row.file or ""
					}
				})

		return sorted(results, key=lambda r: r["score"], reverse=True)[:limit]

	@whitelist()
	def extract_all_files(self):
		"""Extract text from all supported file types (PDF, Excel, CSV)"""
		extracted_count = 0
		error_count = 0
		
		for row in self.documents or []:
			if not row.file:
				continue
				
			try:
				file_ext = row.file.lower().split('.')[-1]
				
				if file_ext == 'pdf':
					if extract_pdf_text_for_row(row):
						extracted_count += 1
						row.is_process = 0
				elif file_ext in ['xlsx', 'xls']:
					if extract_excel_text_for_row(row):
						extracted_count += 1
						row.is_process = 0
				elif file_ext == 'csv':
					if extract_csv_text_for_row(row):
						extracted_count += 1
						row.is_process = 0
			except Exception as e:
				error_count += 1
				frappe.log_error(
					frappe.get_traceback(),
					f"File Extraction Error for Row {row.name}"
				)
				frappe.msgprint(f"Error processing file for row {row.name}: {e}", indicator="red")

		if extracted_count > 0:
			self.save()
			frappe.msgprint(
				f"Successfully extracted text from {extracted_count} file(s). Errors in {error_count} file(s).",
				alert=True,
				indicator="green" if error_count == 0 else "orange"
			)
		else:
			frappe.msgprint("No new files found to extract.", alert=True, indicator="blue")
		
		return {"extracted": extracted_count, "errors": error_count}

	@whitelist()
	def reprocess_all_documents(self):
		"""Mark all documents as unprocessed for re-indexing"""
		for row in self.documents or []:
			row.is_process = 0
		self.save()
		frappe.msgprint(
			"All documents marked for reprocessing. Click 'Upsert to Vector Store' to re-index.",
			alert=True,
			indicator="blue"
		)
	
	@whitelist()
	def get_chunk_stats(self):
		"""Get statistics about chunks that would be created."""
		text_splitter = self.get_text_splitter()
		total_chunks = 0
		total_chars = 0
		
		for row in self.documents or []:
			if row.text_content:
				chunks = text_splitter.split_text(row.text_content)
				total_chunks += len(chunks)
				total_chars += len(row.text_content)
		
		avg_chunk_size = total_chars / total_chunks if total_chunks > 0 else 0
		
		return {
			"total_documents": len(self.documents or []),
			"total_chunks": total_chunks,
			"total_characters": total_chars,
			"avg_chunk_size": int(avg_chunk_size),
			"configured_chunk_size": self.chunk_size or 1000,
			"configured_overlap": self.chunk_overlap or 200,
		}


def _get_provider_api_key(kb: KnowledgeBase):
	try:
		if not kb.provider:
			return None
		provider = frappe.get_doc("LLM Provider", kb.provider)
		return provider.get_password("api_key")
	except Exception:
		return None


def _get_embeddings(kb: KnowledgeBase):
	llm_name = getattr(kb, "embeding_model", None)
	if not llm_name:
		return None
	try:
		llm_doc = frappe.get_doc("LLM", llm_name)
		provider_name = (llm_doc.provider or "")
		api_key = _get_provider_api_key(kb)
		model_name = llm_doc.name
		return create_embedding(provider_name, model=model_name, api_key=api_key)
	except Exception:
		return None


def get_absolute_file_path(file_url: str) -> str:
	"""Converts a Frappe file URL into an absolute server path."""
	if not file_url:
		return ""
	
	clean_url = file_url.lstrip("/")
	site_path = frappe.get_site_path()
	
	if clean_url.startswith("private/"):
		return os.path.join(site_path, clean_url)
	else:
		return os.path.join(site_path, "public", clean_url)


def extract_pdf_text_for_row(row) -> bool:
	"""
	Extract text from PDF file with page-level structure preservation.
	This maintains document structure for better chunking downstream.
	"""
	if not row.file or not row.file.lower().endswith('.pdf'):
		return False
	
	try:
		import pymupdf
	except ImportError:
		frappe.throw(
			"PyMuPDF library not installed. Please run: bench pip install pymupdf",
			title="Library Missing"
		)
	
	file_path = get_absolute_file_path(row.file)
	
	if not os.path.exists(file_path):
		frappe.msgprint(
			f"File not found on server: {row.file}<br>Attempted path: {file_path}",
			alert=True,
			indicator="red",
			title="File Not Found"
		)
		return False
	
	try:
		doc = pymupdf.open(file_path)
		extracted_text = []
		
		for page_num, page in enumerate(doc, 1):
			text = page.get_text("text")
			if text.strip():
				# Preserve page boundaries with clear markers
				extracted_text.append(f"\n\n--- Page {page_num} ---\n\n{text}")
		
		doc.close()
		
		pdf_text = "\n".join(extracted_text)
		
		if not pdf_text.strip():
			frappe.msgprint(
				f"No text could be extracted from PDF: {row.file.split('/')[-1]}", 
				alert=True, 
				indicator="orange"
			)
			return False
		
		# Store with clear separation
		existing_content = row.text_content or ""
		separator = f"\n\n{'='*50}\n[PDF: {row.file.split('/')[-1]}]\n{'='*50}\n"
		
		if separator not in existing_content:
			row.text_content = f"{existing_content}{separator}{pdf_text}".strip()
		
		return True
	
	except Exception as e:
		frappe.log_error(
			f"Error processing PyMuPDF for file {row.file}: {str(e)}", 
			"PDF Extraction Error"
		)
		raise frappe.ValidationError(f"Failed to process PDF '{row.file.split('/')[-1]}': {e}")


def extract_excel_text_for_row(row) -> bool:
	"""Extract text from Excel with sheet-level structure."""
	if not row.file:
		return False
	
	file_ext = row.file.lower().split('.')[-1]
	if file_ext not in ['xlsx', 'xls']:
		return False
	
	try:
		import openpyxl
		import xlrd
	except ImportError:
		frappe.throw(
			"Excel libraries not installed. Please run: bench pip install openpyxl xlrd",
			title="Library Missing"
		)
	
	file_path = get_absolute_file_path(row.file)
	
	if not os.path.exists(file_path):
		frappe.msgprint(
			f"File not found on server: {row.file}<br>Attempted path: {file_path}",
			alert=True,
			indicator="red",
			title="File Not Found"
		)
		return False
	
	try:
		extracted_text = []
		
		if file_ext == 'xlsx':
			wb = openpyxl.load_workbook(file_path, data_only=True)
			for sheet_name in wb.sheetnames:
				sheet = wb[sheet_name]
				extracted_text.append(f"\n\n--- Sheet: {sheet_name} ---\n")
				
				for row_cells in sheet.iter_rows(values_only=True):
					row_data = [str(cell) if cell is not None else "" for cell in row_cells]
					if any(row_data):
						extracted_text.append(" | ".join(row_data))
			wb.close()
		else:
			wb = xlrd.open_workbook(file_path)
			for sheet in wb.sheets():
				extracted_text.append(f"\n\n--- Sheet: {sheet.name} ---\n")
				for row_idx in range(sheet.nrows):
					row_data = [str(cell.value) if cell.value else "" for cell in sheet.row(row_idx)]
					if any(row_data):
						extracted_text.append(" | ".join(row_data))
		
		excel_text = "\n".join(extracted_text)
		
		if not excel_text.strip():
			frappe.msgprint(
				f"No text could be extracted from Excel: {row.file.split('/')[-1]}", 
				alert=True, 
				indicator="orange"
			)
			return False
		
		existing_content = row.text_content or ""
		separator = f"\n\n{'='*50}\n[Excel: {row.file.split('/')[-1]}]\n{'='*50}\n"
		
		if separator not in existing_content:
			row.text_content = f"{existing_content}{separator}{excel_text}".strip()
		
		return True
	
	except Exception as e:
		frappe.log_error(
			f"Error processing Excel file {row.file}: {str(e)}", 
			"Excel Extraction Error"
		)
		raise frappe.ValidationError(f"Failed to process Excel '{row.file.split('/')[-1]}': {e}")


def extract_csv_text_for_row(row) -> bool:
	"""Extract text from CSV file."""
	if not row.file or not row.file.lower().endswith('.csv'):
		return False
	
	file_path = get_absolute_file_path(row.file)
	
	if not os.path.exists(file_path):
		frappe.msgprint(
			f"File not found on server: {row.file}<br>Attempted path: {file_path}",
			alert=True,
			indicator="red",
			title="File Not Found"
		)
		return False
	
	try:
		import csv
		
		extracted_text = []
		extracted_text.append(f"\n\n--- CSV: {row.file.split('/')[-1]} ---\n")
		
		with open(file_path, 'r', encoding='utf-8') as csvfile:
			reader = csv.reader(csvfile)
			for row_data in reader:
				if any(row_data):
					extracted_text.append(" | ".join(row_data))
		
		csv_text = "\n".join(extracted_text)
		
		if not csv_text.strip():
			frappe.msgprint(
				f"No text could be extracted from CSV: {row.file.split('/')[-1]}", 
				alert=True, 
				indicator="orange"
			)
			return False
		
		existing_content = row.text_content or ""
		separator = f"\n\n{'='*50}\n[CSV: {row.file.split('/')[-1]}]\n{'='*50}\n"
		
		if separator not in existing_content:
			row.text_content = f"{existing_content}{separator}{csv_text}".strip()
		
		return True
	
	except Exception as e:
		frappe.log_error(
			f"Error processing CSV file {row.file}: {str(e)}", 
			"CSV Extraction Error"
		)
		raise frappe.ValidationError(f"Failed to process CSV '{row.file.split('/')[-1]}': {e}")