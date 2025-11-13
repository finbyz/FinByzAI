# Copyright (c) 2025, Finbyz Tech Pvt Ltd and contributors
# For license information, please see license.txt

import frappe
from frappe.model.document import Document
from frappe.utils import cstr
import os


class KnowledgeDocument(Document):
	"""Knowledge Document child table row with auto text extraction from multiple file types"""
	
	# Supported file extensions
	SUPPORTED_EXTENSIONS = ['.pdf', '.txt', '.csv', '.xlsx', '.xls']
	
	def validate(self):
		"""
		On validate, trigger text extraction if the file has changed
		and mark for reprocessing if content is modified.
		"""
		# Validate required fields exist
		self._validate_required_fields()
		
		# Only extract if file is new or has been changed
		if self.has_value_changed("file") and self.file:
			self.auto_extract_content()
		
		# Mark as not processed if content changed, ensuring it gets re-indexed
		if self.has_value_changed("text_content") or self.has_value_changed("file"):
			self.is_process = False
			# Optional: Add timestamp for tracking
			self.last_extraction_attempt = frappe.utils.now()
	
	def _validate_required_fields(self):
		"""Ensure required fields exist in the doctype"""
		required_fields = ['text_content', 'is_process', 'file']
		missing_fields = [f for f in required_fields if not hasattr(self, f)]
		
		if missing_fields:
			frappe.throw(
				f"Missing required fields in Knowledge Document: {', '.join(missing_fields)}"
			)
	
	def auto_extract_content(self):
		"""Automatically extract text from supported file types on save."""
		if not self.file:
			return
		
		# Get file extension
		file_ext = self._get_file_extension()
		
		# Check if file type is supported
		if file_ext not in self.SUPPORTED_EXTENSIONS:
			# Silent return for unsupported files
			return
		
		# Validate file exists in the system
		if not self._validate_file_exists():
			return
		
		# Route to appropriate extraction method based on file type
		try:
			extraction_method = self._get_extraction_method(file_ext)
			
			if extraction_method:
				extraction_success = extraction_method()
				
				if extraction_success:
					file_name = self.file.split('/')[-1]
					frappe.msgprint(
						f"Text successfully extracted from: <b>{file_name}</b>",
						alert=True,
						indicator="green",
						title=f"{file_ext.upper()} Processed"
					)
					
					# Log successful extraction
					frappe.log(
						f"Content extracted successfully from {file_ext} file: {file_name}"
					)
				else:
					self._handle_extraction_failure("Extraction function returned no result")
					
		except Exception as e:
			# Log the error and show a message but DO NOT block the save
			self._handle_extraction_failure(str(e))
	
	def _get_file_extension(self):
		"""Get the file extension in lowercase"""
		if not self.file:
			return None
		return os.path.splitext(self.file.lower())[1]
	
	def _get_extraction_method(self, file_ext):
		"""Return the appropriate extraction method based on file extension"""
		extraction_methods = {
			'.pdf': self._extract_pdf,
			'.txt': self._extract_text,
			'.csv': self._extract_csv,
			'.xlsx': self._extract_excel,
			'.xls': self._extract_excel,
		}
		return extraction_methods.get(file_ext)
	
	def _extract_pdf(self):
		"""Extract text from PDF files"""
		try:
			from finbyzai.ai.doctype.knowledge_base.knowledge_base import extract_pdf_text_for_row
			return extract_pdf_text_for_row(self)
		except ImportError as e:
			frappe.log_error(
				f"Could not import PDF extraction function: {e}",
				"PDF Extraction Import Error"
			)
			frappe.msgprint(
				"PDF extraction function is not available. Please contact your system administrator.",
				alert=True,
				indicator="red",
				title="Configuration Error"
			)
			return False
	
	def _extract_text(self):
		"""Extract content from plain text files"""
		try:
			# Get the file path
			file_path = frappe.get_site_path('public', self.file.lstrip('/'))
			
			# Read the text file
			with open(file_path, 'r', encoding='utf-8') as f:
				text_content = f.read()
			
			# Store in text_content field
			self.text_content = text_content
			return True
			
		except UnicodeDecodeError:
			# Try with different encoding if UTF-8 fails
			try:
				with open(file_path, 'r', encoding='latin-1') as f:
					text_content = f.read()
				self.text_content = text_content
				return True
			except Exception as e:
				frappe.log_error(f"Failed to read text file with alternate encoding: {e}")
				return False
				
		except Exception as e:
			frappe.log_error(f"Error reading text file: {e}", "Text File Extraction Error")
			return False
	
	def _extract_csv(self):
		"""Extract content from CSV files"""
		try:
			import csv
			
			# Get the file path
			file_path = frappe.get_site_path('public', self.file.lstrip('/'))
			
			# Read CSV and convert to readable text format
			rows = []
			with open(file_path, 'r', encoding='utf-8') as f:
				csv_reader = csv.reader(f)
				headers = next(csv_reader, None)
				
				if headers:
					rows.append("Headers: " + " | ".join(headers))
					rows.append("-" * 50)
				
				# Read all rows
				for row_num, row in enumerate(csv_reader, 1):
					if row_num <= 1000:  # Limit to first 1000 rows to avoid too much data
						rows.append(" | ".join(str(cell) for cell in row))
					else:
						rows.append(f"\n... (truncated, showing first 1000 rows)")
						break
			
			# Store formatted text
			self.text_content = "\n".join(rows)
			return True
			
		except UnicodeDecodeError:
			# Try with different encoding
			try:
				rows = []
				with open(file_path, 'r', encoding='latin-1') as f:
					csv_reader = csv.reader(f)
					headers = next(csv_reader, None)
					
					if headers:
						rows.append("Headers: " + " | ".join(headers))
						rows.append("-" * 50)
					
					for row_num, row in enumerate(csv_reader, 1):
						if row_num <= 1000:
							rows.append(" | ".join(str(cell) for cell in row))
						else:
							rows.append(f"\n... (truncated, showing first 1000 rows)")
							break
				
				self.text_content = "\n".join(rows)
				return True
			except Exception as e:
				frappe.log_error(f"Failed to read CSV with alternate encoding: {e}")
				return False
				
		except Exception as e:
			frappe.log_error(f"Error reading CSV file: {e}", "CSV File Extraction Error")
			return False
	
	def _extract_excel(self):
		"""Extract content from Excel files (.xlsx, .xls)"""
		try:
			import openpyxl
			from openpyxl import load_workbook
			
			# Get the file path
			file_path = frappe.get_site_path('public', self.file.lstrip('/'))
			
			# Load the workbook
			workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
			
			all_sheets_text = []
			
			# Process each sheet
			for sheet_name in workbook.sheetnames[:10]:  # Limit to first 10 sheets
				sheet = workbook[sheet_name]
				
				all_sheets_text.append(f"\n{'='*50}")
				all_sheets_text.append(f"Sheet: {sheet_name}")
				all_sheets_text.append(f"{'='*50}\n")
				
				rows = []
				row_count = 0
				
				for row in sheet.iter_rows(values_only=True):
					if row_count >= 1000:  # Limit rows per sheet
						rows.append("\n... (truncated, showing first 1000 rows)")
						break
					
					# Filter out None values and convert to string
					row_data = [str(cell) if cell is not None else "" for cell in row]
					
					# Only add non-empty rows
					if any(cell.strip() for cell in row_data):
						rows.append(" | ".join(row_data))
						row_count += 1
				
				all_sheets_text.extend(rows)
			
			# Store formatted text
			self.text_content = "\n".join(all_sheets_text)
			workbook.close()
			return True
			
		except ImportError:
			frappe.log_error(
				"openpyxl library not installed. Please install it to extract Excel files.",
				"Excel Extraction Error"
			)
			frappe.msgprint(
				"Excel extraction requires openpyxl library. Please contact your system administrator.",
				alert=True,
				indicator="orange",
				title="Library Missing"
			)
			return False
			
		except Exception as e:
			frappe.log_error(f"Error reading Excel file: {e}", "Excel File Extraction Error")
			return False
	
	def _validate_file_exists(self):
		"""Check if the file exists in the system"""
		try:
			# Check if file exists in File doctype
			file_exists = frappe.db.exists("File", {"file_url": self.file})
			
			if not file_exists:
				frappe.msgprint(
					f"File not found in system: {self.file.split('/')[-1]}",
					alert=True,
					indicator="orange",
					title="File Not Found"
				)
				return False
			
			return True
			
		except Exception as e:
			frappe.log_error(
				f"Error validating file existence: {self.file}\n{e}",
				"File Validation Error"
			)
			return False
	
	def _handle_extraction_failure(self, error_message):
		"""Centralized handler for extraction failures"""
		file_name = self.file.split('/')[-1] if self.file else "Unknown"
		file_ext = self._get_file_extension()
		
		# Log the error
		frappe.log_error(
			f"Auto-extraction failed for {file_ext} file: {self.file}\n{error_message}",
			f"{file_ext.upper()} Auto-Extraction Error"
		)
		
		# User-friendly message
		frappe.msgprint(
			f"Could not automatically extract text from '<b>{file_name}</b>'. "
			"You can try again using the 'Extract Text' button on the Knowledge Base.",
			alert=True,
			indicator="orange",
			title="Extraction Warning"
		)
	
	def manual_extract(self):
		"""
		Public method to manually trigger text extraction.
		Can be called from a button or API endpoint.
		"""
		if not self.file:
			frappe.throw("No file attached to extract text from")
		
		file_ext = self._get_file_extension()
		
		if file_ext not in self.SUPPORTED_EXTENSIONS:
			frappe.throw(
				f"Unsupported file type: {file_ext}. "
				f"Supported types: {', '.join(self.SUPPORTED_EXTENSIONS)}"
			)
		
		# Force extraction regardless of previous status
		self.auto_extract_content()
		
		# Save the document to persist extracted text
		self.save(ignore_permissions=True)
		
		return {
			"success": True,
			"message": f"Text extracted from {self.file.split('/')[-1]}",
			"file_type": file_ext
		}
