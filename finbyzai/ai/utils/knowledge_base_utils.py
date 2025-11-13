# Copyright (c) 2025, Finbyz Tech Pvt Ltd and contributors
# For license information, please see license.txt

"""
Utility functions for extracting text content from various file formats and web links.
These functions are decoupled from the Knowledge Document class for reusability.
"""

import fitz
import frappe
import os
import csv
import requests
from urllib.parse import urlparse
from typing import Tuple, Optional, Dict, Any
from bs4 import BeautifulSoup


# Supported file extensions
SUPPORTED_FILE_EXTENSIONS = ['.pdf', '.txt', '.csv', '.xlsx', '.xls']
SUPPORTED_WEB_CONTENT_TYPES = ['text/html', 'text/plain', 'application/pdf']


def get_file_extension(file_path: str) -> Optional[str]:
    """
    Get the file extension in lowercase.
    
    Args:
        file_path: Path or URL of the file
        
    Returns:
        Lowercase file extension including the dot (e.g., '.pdf'), or None if no path
    """
    if not file_path:
        return None
    return os.path.splitext(file_path.lower())[1]


def is_supported_file_type(file_path: str) -> bool:
    """
    Check if the file type is supported for text extraction.
    
    Args:
        file_path: Path or URL of the file
        
    Returns:
        True if file type is supported, False otherwise
    """
    ext = get_file_extension(file_path)
    return ext in SUPPORTED_FILE_EXTENSIONS


def is_web_url(file_path: str) -> bool:
    """
    Check if the given path is a web URL.
    
    Args:
        file_path: Path or URL to check
        
    Returns:
        True if it's a web URL, False otherwise
    """
    if not file_path:
        return False
    
    try:
        result = urlparse(file_path)
        return all([result.scheme in ['http', 'https'], result.netloc])
    except Exception:
        return False


def get_file_path_from_url(file_url: str) -> str:
    """
    Convert file URL to absolute file path.
    
    Args:
        file_url: File URL (e.g., '/files/document.pdf')
        
    Returns:
        Absolute file path
    """
    if file_url.startswith("/private/files/"):
        return os.path.join(
            frappe.get_site_path("private", "files"), 
            os.path.basename(file_url)
        )
    else:
        # Remove leading slash if present
        file_url = file_url.lstrip('/')
        return frappe.get_site_path('public', file_url)


def validate_file_exists(file_url: str) -> bool:
    """
    Check if the file exists in the Frappe File doctype.
    
    Args:
        file_url: File URL to validate
        
    Returns:
        True if file exists, False otherwise
    """
    try:
        return frappe.db.exists("File", {"file_url": file_url})
    except Exception as e:
        frappe.log_error(
            f"Error validating file existence: {file_url}\n{e}",
            "File Validation Error"
        )
        return False


def extract_text_from_file(file_path: str, encoding: str = 'utf-8') -> Tuple[bool, Optional[str], Optional[str]]:
    """
    Extract text content from a plain text file.
    
    Args:
        file_path: Absolute path to the text file
        encoding: Text encoding (default: 'utf-8')
        
    Returns:
        Tuple of (success, content, error_message)
    """
    try:
        with open(file_path, 'r', encoding=encoding) as f:
            content = f.read()
        return True, content, None
        
    except UnicodeDecodeError:
        # Try with alternate encoding
        try:
            with open(file_path, 'r', encoding='latin-1') as f:
                content = f.read()
            return True, content, None
        except Exception as e:
            return False, None, f"Failed with alternate encoding: {str(e)}"
            
    except Exception as e:
        return False, None, str(e)


def extract_csv_to_text(file_path: str, max_rows: int = 1000, encoding: str = 'utf-8') -> Tuple[bool, Optional[str], Optional[str]]:
    """
    Extract CSV content and format as readable text.
    
    Args:
        file_path: Absolute path to the CSV file
        max_rows: Maximum number of rows to extract (default: 1000)
        encoding: Text encoding (default: 'utf-8')
        
    Returns:
        Tuple of (success, content, error_message)
    """
    try:
        rows = []
        with open(file_path, 'r', encoding=encoding) as f:
            csv_reader = csv.reader(f)
            headers = next(csv_reader, None)
            
            if headers:
                rows.append("Headers: " + " | ".join(headers))
                rows.append("-" * 50)
            
            for row_num, row in enumerate(csv_reader, 1):
                if row_num <= max_rows:
                    rows.append(" | ".join(str(cell) for cell in row))
                else:
                    rows.append(f"\n... (truncated, showing first {max_rows} rows)")
                    break
        
        content = "\n".join(rows)
        return True, content, None
        
    except UnicodeDecodeError:
        # Try with alternate encoding
        try:
            rows = []
            with open(file_path, 'r', encoding='latin-1') as f:
                csv_reader = csv.reader(f)
                headers = next(csv_reader, None)
                
                if headers:
                    rows.append("Headers: " + " | ".join(headers))
                    rows.append("-" * 50)
                
                for row_num, row in enumerate(csv_reader, 1):
                    if row_num <= max_rows:
                        rows.append(" | ".join(str(cell) for cell in row))
                    else:
                        rows.append(f"\n... (truncated, showing first {max_rows} rows)")
                        break
            
            content = "\n".join(rows)
            return True, content, None
        except Exception as e:
            return False, None, f"Failed with alternate encoding: {str(e)}"
            
    except Exception as e:
        return False, None, str(e)


def extract_excel_to_text(file_path: str, max_sheets: int = 10, max_rows_per_sheet: int = 1000) -> Tuple[bool, Optional[str], Optional[str]]:
    """
    Extract Excel content and format as readable text.
    
    Args:
        file_path: Absolute path to the Excel file
        max_sheets: Maximum number of sheets to process (default: 10)
        max_rows_per_sheet: Maximum rows per sheet (default: 1000)
        
    Returns:
        Tuple of (success, content, error_message)
    """
    try:
        from openpyxl import load_workbook
        
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
        all_sheets_text = []
        
        # Process each sheet
        for sheet_name in workbook.sheetnames[:max_sheets]:
            sheet = workbook[sheet_name]
            
            all_sheets_text.append(f"\n{'='*50}")
            all_sheets_text.append(f"Sheet: {sheet_name}")
            all_sheets_text.append(f"{'='*50}\n")
            
            rows = []
            row_count = 0
            
            for row in sheet.iter_rows(values_only=True):
                if row_count >= max_rows_per_sheet:
                    rows.append(f"\n... (truncated, showing first {max_rows_per_sheet} rows)")
                    break
                
                # Filter out None values and convert to string
                row_data = [str(cell) if cell is not None else "" for cell in row]
                
                # Only add non-empty rows
                if any(cell.strip() for cell in row_data):
                    rows.append(" | ".join(row_data))
                    row_count += 1
            
            all_sheets_text.extend(rows)
        
        content = "\n".join(all_sheets_text)
        workbook.close()
        return True, content, None
        
    except ImportError:
        return False, None, "openpyxl library not installed. Please install it to extract Excel files."
        
    except Exception as e:
        return False, None, str(e)


def extract_text_from_web_url(url: str, timeout: int = 10) -> Tuple[bool, Optional[str], Optional[str]]:
    """
    Extract text content from a web URL.
    
    Args:
        url: Web URL to extract content from
        timeout: Request timeout in seconds (default: 10)
        
    Returns:
        Tuple of (success, content, error_message)
    """
    try:
        headers = {
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/91.0.4472.124 Safari/537.36'
        }
        
        response = requests.get(url, headers=headers, timeout=timeout, allow_redirects=True)
        response.raise_for_status()
        
        content_type = response.headers.get('content-type', '').lower()
        
        # Handle HTML content
        if 'text/html' in content_type:
            soup = BeautifulSoup(response.content, 'html.parser')
            
            # Remove script and style elements
            for script in soup(["script", "style"]):
                script.decompose()
            
            # Get text and clean it up
            text = soup.get_text()
            
            # Clean up extra whitespace
            lines = (line.strip() for line in text.splitlines())
            chunks = (phrase.strip() for line in lines for phrase in line.split("  "))
            text = ' '.join(chunk for chunk in chunks if chunk)
            
            return True, text, None
            
        # Handle plain text content
        elif 'text/plain' in content_type:
            text = response.text
            return True, text, None
            
        else:
            return False, None, f"Unsupported web content type: {content_type}"
            
    except requests.exceptions.RequestException as e:
        return False, None, f"HTTP request failed: {str(e)}"
    except Exception as e:
        return False, None, f"Web content extraction failed: {str(e)}"


def extract_pdf_to_text(file_path: str) -> Tuple[bool, Optional[str], Optional[str]]:
    """
    Extract text content from PDF using PyMuPDF (recommended).
    
    Args:
        file_path: Absolute path to the PDF file
        
    Returns:
        Tuple of (success, content, error_message)
    """
    try:
        
        doc = None
        try:
            doc = fitz.open(file_path)
            text_content = []
            
            for page_num in range(len(doc)):
                page = doc[page_num]
                text = page.get_text()
                
                if text.strip():
                    text_content.append(f"--- Page {page_num + 1} ---")
                    text_content.append(text)
            
            content = "\n".join(text_content)
            
            if not content.strip():
                return False, None, "No extractable text found in PDF"
                
            return True, content, None
            
        except Exception as e:
            return False, None, f"PDF extraction failed: {str(e)}"
        finally:
            if doc:
                doc.close()
                
    except ImportError:
        return False, None, "PyMuPDF library not installed. Please install it with: pip install PyMuPDF"




def extract_text_from_source(source: str, source_type: Optional[str] = None) -> Dict[str, Any]:
    """
    Universal function to extract text from files, URLs, or web links.
    
    Args:
        source: File URL, file path, or web URL
        source_type: Optional type hint ('file', 'url', or auto-detected)
        
    Returns:
        Dictionary containing:
            - success (bool): Whether extraction succeeded
            - content (str | None): Extracted text content
            - error (str | None): Error message if failed
            - source_type (str): Detected source type ('file' or 'web_url')
            - file_type (str | None): File extension if applicable
    """
    # Auto-detect source type if not provided
    if source_type is None:
        if is_web_url(source):
            source_type = 'web_url'
        else:
            source_type = 'file'
    
    if source_type == 'web_url':
        # Extract from web URL
        success, content, error = extract_text_from_web_url(source)
        return {
            'success': success,
            'content': content,
            'error': error,
            'source_type': 'web_url',
            'file_type': None
        }
    
    else:  # source_type == 'file'
        # Validate file exists (for Frappe file URLs)
        if source.startswith('/') and not validate_file_exists(source):
            return {
                'success': False,
                'content': None,
                'error': 'File not found in system',
                'source_type': 'file',
                'file_type': None
            }
        
        # Get file extension
        file_type = get_file_extension(source)
        
        # Check if supported
        if file_type not in SUPPORTED_FILE_EXTENSIONS:
            return {
                'success': False,
                'content': None,
                'error': f'Unsupported file type: {file_type}',
                'source_type': 'file',
                'file_type': file_type
            }
        
        # Get absolute file path
        file_path = get_file_path_from_url(source)
        
        # Route to appropriate extraction function
        if file_type == '.txt':
            success, content, error = extract_text_from_file(file_path)
        elif file_type == '.csv':
            success, content, error = extract_csv_to_text(file_path)
        elif file_type in ['.xlsx', '.xls']:
            success, content, error = extract_excel_to_text(file_path)
        elif file_type == '.pdf':
            # PDF extraction requires external function
            success, content, error = extract_pdf_to_text(file_path)
            
        else:
            return {
                'success': False,
                'content': None,
                'error': f'No extraction handler for {file_type}',
                'source_type': 'file',
                'file_type': file_type
            }
        
        return {
            'success': success,
            'content': content,
            'error': error,
            'source_type': 'file',
            'file_type': file_type
        }


def extract_text_from_url(file_url: str, file_type: Optional[str] = None) -> Dict[str, Any]:
    """
    Legacy function - maintained for backward compatibility.
    Use extract_text_from_source for new code.
    
    Args:
        file_url: File URL from Frappe
        file_type: File extension override (e.g., '.pdf')
        
    Returns:
        Dictionary containing:
            - success (bool): Whether extraction succeeded
            - content (str | None): Extracted text content
            - error (str | None): Error message if failed
            - file_type (str | None): Detected or provided file type
    """
    return extract_text_from_source(file_url, 'file')


